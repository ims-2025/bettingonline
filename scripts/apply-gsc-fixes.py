#!/usr/bin/env python3
"""
Phase 2 GSC fixes:

1. Add redirects to vercel.json for the 74 not-found (404) URLs
   - WordPress /tag/*, /category/*, /author/*/page/* → 301 to /
   - Broken /us/reviews/fanduel/, /sports/reviews/fanduel/ etc. → 301 to /reviews/*
   - Misc bad paths → 301 to closest match

2. Add noindex meta to 273 crawled-not-indexed pages
   - Skip URLs that don't correspond to a file on disk (they'll be
     handled by the 404 redirect layer)
   - For legacy blog slugs and thin cluster pages: inject noindex meta
   - Skip URLs that ARE currently earning impressions (protect ranking assets)

3. Remove noindexed pages from sitemap.xml so we don't confuse Google
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
VERCEL = ROOT / "vercel.json"
SITEMAP = ROOT / "sitemap.xml"

# Import openpyxl only where needed
import openpyxl

# ---------------------------------------------------------------------------
# Load GSC data
# ---------------------------------------------------------------------------
UPLOADS = Path("/sessions/exciting-trusting-wozniak/mnt/uploads")


def load_urls(filename: str) -> list[str]:
    wb = openpyxl.load_workbook(UPLOADS / filename)
    ws = wb["Table"]
    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            urls.append(row[0].replace("https://www.bettingonline.org", ""))
    return urls


NOT_FOUND = load_urls("https___www.bettingonline.org_-Coverage-Drilldown-2026-07-16.xlsx")
CRAWLED_NOT_INDEXED = load_urls("https___www.bettingonline.org_-Coverage-Drilldown-2026-07-16 (1).xlsx")

print(f"Loaded {len(NOT_FOUND)} not-found URLs")
print(f"Loaded {len(CRAWLED_NOT_INDEXED)} crawled-not-indexed URLs")

# ---------------------------------------------------------------------------
# 1. Build redirect list for vercel.json
# ---------------------------------------------------------------------------
new_redirects = []
seen = set()


def add_redirect(source: str, destination: str, permanent: bool = True):
    if source in seen:
        return
    seen.add(source)
    new_redirects.append({"source": source, "destination": destination, "permanent": permanent})


# Wildcard patterns first — these cover most tag/category/author URLs
# Redirect all remaining /tag/* and /category/* to the homepage. Google will
# eventually drop them from the "not found" report on next crawl.
add_redirect("/tag/:path*", "/", True)
add_redirect("/category/:path*", "/", True)
add_redirect("/author/:name/page/:page", "/about/", True)

# Broken nested review paths — someone linked to /us/reviews/fanduel/
# instead of /reviews/fanduel/. Fix at the URL layer.
add_redirect("/us/reviews/:slug/", "/reviews/:slug/", True)
add_redirect("/sports/reviews/:slug/", "/reviews/:slug/", True)
add_redirect("/fighting/reviews/:slug/", "/reviews/:slug/", True)
add_redirect("/casino/reviews/:slug/", "/reviews/:slug/", True)
add_redirect("/poker/reviews/:slug/", "/reviews/:slug/", True)

# Odd query-string 404
add_redirect("/odds/?apiKey", "/odds/", True)

# Orphan legacy slug
add_redirect("/kentucky-2/", "/us/", True)

# Malformed URLs like /us//1000 and /us/florida//1000 that Google indexed
# from bad internal links — send them to the state hub.
add_redirect("/us//:path*", "/us/", True)
add_redirect("/us/:slug/:path*", "/us/:slug/", True)

# Misspelled review URLs from the 273 crawled-not-indexed list
add_redirect("/reviews/fanduel-sportsbooks/", "/reviews/fanduel-sportsbook/", True)
add_redirect("/reviews/fanduel-sportsbooks-review-2025/", "/reviews/fanduel-sportsbook/", True)
add_redirect("/reviews/draftkings-sportsbook-review-2025-real-money-testing-of-every-feature/", "/reviews/draftkings-sportsbook/", True)
add_redirect("/us/arkansas-sports-betting/", "/us/arkansas/", True)


# ---------------------------------------------------------------------------
# 2. Patch vercel.json
# ---------------------------------------------------------------------------
config = json.loads(VERCEL.read_text())
existing_sources = {r["source"] for r in config.get("redirects", [])}
added = 0
for r in new_redirects:
    if r["source"] not in existing_sources:
        config["redirects"].append(r)
        added += 1
VERCEL.write_text(json.dumps(config, indent=2) + "\n")
print(f"Added {added} new redirects to vercel.json")


# ---------------------------------------------------------------------------
# 3. Add noindex to files corresponding to crawled-not-indexed URLs
#
# Strategy: for each URL in the 273 list, if a file exists on disk, add
# <meta name="robots" content="noindex, follow"> to the <head>. This tells
# Google "we agree, don't index this" — cleans up the crawl budget and
# removes the low-quality signal.
#
# PROTECTED — do NOT noindex these even if listed (they're earning impressions):
# We cross-reference the GSC Pages export to identify URLs earning impressions
# and skip them.
# ---------------------------------------------------------------------------

# Load the impression data to protect earning URLs
wb_perf = openpyxl.load_workbook(UPLOADS / "https___www.bettingonline.org_-Performance-on-Search-2026-07-16.xlsx")
ws_pages = wb_perf["Pages"]
earning_urls = set()
for row in ws_pages.iter_rows(min_row=2, values_only=True):
    if row[0] and (row[2] or 0) > 0:  # any impressions at all
        u = row[0].replace("https://www.bettingonline.org", "").rstrip("/")
        earning_urls.add(u)
        earning_urls.add(u + "/")

print(f"Protecting {len(earning_urls)//2} URLs currently earning GSC impressions")


NOINDEX_META = '  <meta name="robots" content="noindex, follow">\n'

def url_to_file(url: str) -> Path | None:
    """Map a URL back to a local file path."""
    url = url.strip("/")
    if not url:
        return ROOT / "index.html"
    # try /path/index.html first
    p = ROOT / url / "index.html"
    if p.exists():
        return p
    # try /path.html
    p = ROOT / f"{url}.html"
    if p.exists():
        return p
    return None


noindexed_count = 0
skipped_earning = 0
skipped_missing = 0
noindexed_urls = set()

for url in CRAWLED_NOT_INDEXED:
    # Skip if this URL is currently earning impressions
    if url.rstrip("/") in earning_urls or url in earning_urls:
        skipped_earning += 1
        continue

    path = url_to_file(url)
    if not path:
        skipped_missing += 1
        continue

    text = path.read_text()

    # Already has noindex? skip
    if re.search(r'<meta[^>]+robots[^>]+noindex', text, re.I):
        continue

    # Inject noindex into <head>. Prefer right after <meta charset>.
    if '<meta charset=' in text:
        text2 = re.sub(
            r'(<meta charset="UTF-8">)',
            r'\1\n' + NOINDEX_META.strip(),
            text, count=1
        )
    else:
        # Fall back — insert after <head>
        text2 = text.replace('<head>', '<head>\n' + NOINDEX_META.strip(), 1)

    if text2 != text:
        path.write_text(text2)
        noindexed_count += 1
        noindexed_urls.add("https://www.bettingonline.org" + url)

print(f"Added noindex to {noindexed_count} pages")
print(f"Skipped {skipped_earning} pages that are earning impressions (protected)")
print(f"Skipped {skipped_missing} URLs that don't have a corresponding file (will be handled by redirects)")

# ---------------------------------------------------------------------------
# 4. Remove noindexed URLs from sitemap
# ---------------------------------------------------------------------------
if noindexed_urls:
    sm_text = SITEMAP.read_text()
    urlset = re.split(r'(<url>.*?</url>\s*)', sm_text, flags=re.S)
    kept = []
    dropped = 0
    for block in urlset:
        if block.strip().startswith("<url>"):
            m = re.search(r'<loc>([^<]+)</loc>', block)
            if m and m.group(1) in noindexed_urls:
                dropped += 1
                continue
        kept.append(block)
    SITEMAP.write_text("".join(kept))
    print(f"Removed {dropped} noindexed URLs from sitemap.xml")


# ---------------------------------------------------------------------------
# Commit + push
# ---------------------------------------------------------------------------
subprocess.run(["git", "-C", str(ROOT), "add", "-A"], check=True)
